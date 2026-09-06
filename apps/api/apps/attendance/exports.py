"""Rendering §13's report rows into the three formats §6 names.

One writer per format over one row shape, deliberately: `tasks.build_report_rows`
produces the rows, and everything here only decides how they are laid out. That
is what makes adding a format a formatter rather than a query, and it is why the
numbers in a CSV, a spreadsheet and a printed PDF cannot drift apart — a school
that reconciles an exported register against a printed one is checking the
school's data, not this module's three code paths.

Both heavy imports are **lazy**, and for different reasons. `openpyxl` is cheap
but only ever needed here. `weasyprint` needs system libraries that only
`.github/workflows/api.yml`'s `test` job installs, so a module-level import
would break `manage.py` everywhere else — the convention every WeasyPrint caller
in this repo already follows.
"""

from __future__ import annotations

import csv
import datetime
import io
from html import escape

# `text/csv` matches the upload-purpose registry's declared MIME for
# `attendance.report-export`; the other two are added there alongside.
FORMATS = {
    "csv": ("text/csv", "csv"),
    "xlsx": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xlsx",
    ),
    "pdf": ("application/pdf", "pdf"),
}

# A PDF is paginated and read at arm's length, so it gets a hard cap that CSV and
# XLSX do not: a 40,000-row register rendered to PDF is hundreds of pages nobody
# opens and enough memory to matter, and WeasyPrint is the slowest thing in this
# module by a wide margin. The caller is told, rather than being handed a
# truncated document that looks complete.
PDF_ROW_LIMIT = 2000


class ReportTooLargeForFormat(Exception):
    """Raised when a format cannot honestly represent this many rows."""


def render(rows: list[dict], *, fmt: str, title: str) -> tuple[bytes, str, str]:
    """Render `rows` as `fmt`. Returns `(bytes, mime_type, extension)`."""
    if fmt not in FORMATS:
        raise ValueError(f"Unknown export format {fmt!r}")
    mime_type, extension = FORMATS[fmt]

    if fmt == "csv":
        return _csv(rows), mime_type, extension
    if fmt == "xlsx":
        return _xlsx(rows, title=title), mime_type, extension
    return _pdf(rows, title=title), mime_type, extension


def _headers(rows: list[dict]) -> list[str]:
    """Column order taken from the first row.

    Every row in a report comes from one `.values()` call, so the keys are
    uniform and insertion-ordered — the order `reports.py` wrote them in, which
    is the order a reader expects rather than an alphabetical scramble.
    """
    return list(rows[0]) if rows else []


def _cell(value: object) -> str:
    """A display string for one cell.

    Dates are ISO rather than locale-formatted: an export is read by a
    spreadsheet at least as often as by a person, and a locale-formatted date is
    the classic way a column of dates becomes a column of text.
    """
    if value is None:
        return ""
    if isinstance(value, datetime.date | datetime.time):
        return value.isoformat()
    return str(value)


# OWASP's CSV-injection set, plus the two whitespace characters Excel strips
# before parsing — a cell starting with a tab or carriage return followed by `=`
# is still a formula.
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def _spreadsheet_safe(text: str) -> str:
    """Neutralise a value a spreadsheet would otherwise execute as a formula.

    `remarks` is free text a teacher types at the register, so a remark reading
    `=HYPERLINK("http://evil","Click")` executes the moment someone opens the
    export in Excel, Sheets or LibreOffice — a client-side execution vector that
    starts inside our own data and needs no other flaw.

    The leading apostrophe is the standard mitigation: every major spreadsheet
    reads it as "the rest is text" and hides it in the cell. It does show in a
    raw CSV opened in a text editor, which is the accepted cost — a visible
    apostrophe is a far smaller problem than a live formula.

    **Not applied to the PDF**, which escapes HTML instead: there is no formula
    engine in a PDF, and a stray apostrophe in a printed register would be a
    defect with nothing to justify it.
    """
    return f"'{text}" if text.startswith(_FORMULA_TRIGGERS) else text


def _csv(rows: list[dict]) -> bytes:
    buffer = io.StringIO()
    if not rows:
        # An empty file is indistinguishable from a failed export when someone
        # opens it, so it says which it is.
        buffer.write("no rows matched this report\n")
        return buffer.getvalue().encode()

    writer = csv.DictWriter(buffer, fieldnames=_headers(rows))
    writer.writeheader()
    writer.writerows(
        {key: _spreadsheet_safe(_cell(value)) for key, value in row.items()} for row in rows
    )
    return buffer.getvalue().encode()


def _xlsx(rows: list[dict], *, title: str) -> bytes:
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Excel refuses a sheet name over 31 characters or containing []:*?/\, and
    # fails the whole save rather than truncating — so the title is sanitised
    # here instead of being trusted.
    sheet.title = _sheet_name(title)

    headers = _headers(rows)
    if not headers:
        sheet["A1"] = "no rows matched this report"
        return _workbook_bytes(workbook)

    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    # Freeze the header so a scrolled register still says which column is which
    # — the one affordance that makes a wide export usable at all.
    sheet.freeze_panes = "A2"

    for row in rows:
        # openpyxl writes a string beginning `=` as a *formula*, so this is
        # not merely defence against the reader's spreadsheet — it is what
        # stops us writing one ourselves.
        sheet.append([_spreadsheet_safe(_cell(row.get(header))) for header in headers])

    for index, header in enumerate(headers, start=1):
        widest = max((len(_cell(row.get(header))) for row in rows), default=0)
        sheet.column_dimensions[get_column_letter(index)].width = min(
            max(len(header), widest) + 2, 40
        )

    return _workbook_bytes(workbook)


def _workbook_bytes(workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _sheet_name(title: str) -> str:
    for character in "[]:*?/\\":
        title = title.replace(character, "-")
    return title[:31] or "Report"


def _pdf(rows: list[dict], *, title: str) -> bytes:
    from weasyprint import HTML

    if len(rows) > PDF_ROW_LIMIT:
        raise ReportTooLargeForFormat(
            f"{len(rows)} rows is too many to print; export CSV or XLSX instead "
            f"(the PDF limit is {PDF_ROW_LIMIT})."
        )

    headers = _headers(rows)
    # **Every value escaped.** These are student and staff names straight out of
    # the database, and a name containing `&` or `<` would otherwise break the
    # table silently — a register missing a row is worse than one that fails.
    header_html = "".join(f"<th>{escape(_readable(header))}</th>" for header in headers)
    body_html = "".join(
        "<tr>"
        + "".join(f"<td>{escape(_cell(row.get(header)))}</td>" for header in headers)
        + "</tr>"
        for row in rows
    )
    if not rows:
        body_html = "<tr><td>no rows matched this report</td></tr>"
        header_html = "<th>Result</th>"

    document = f"""
    <html>
      <head>
        <meta charset="utf-8" />
        <style>
          @page {{ size: A4 landscape; margin: 12mm; }}
          body {{ font-family: sans-serif; font-size: 9pt; }}
          h1 {{ font-size: 14pt; margin: 0 0 2mm; }}
          .generated {{ color: #555; font-size: 8pt; margin: 0 0 4mm; }}
          table {{ border-collapse: collapse; width: 100%; }}
          th, td {{ border: 0.4pt solid #999; padding: 1.2mm 1.8mm; text-align: left; }}
          /* Repeat the header on every page: a printed register whose columns
             are only labelled on page one is unreadable from page two. */
          thead {{ display: table-header-group; }}
          tr {{ page-break-inside: avoid; }}
        </style>
      </head>
      <body>
        <h1>{escape(title)}</h1>
        <p class="generated">Generated {datetime.date.today().isoformat()}</p>
        <table>
          <thead><tr>{header_html}</tr></thead>
          <tbody>{body_html}</tbody>
        </table>
      </body>
    </html>
    """
    return HTML(string=document).write_pdf()


def _readable(header: str) -> str:
    """`total_late_minutes` -> `Total late minutes`, for a printed heading.

    Only for the PDF: CSV and XLSX keep the machine names, because something
    downstream parses those and a prettified header is a broken import.
    """
    return header.replace("_", " ").capitalize()
