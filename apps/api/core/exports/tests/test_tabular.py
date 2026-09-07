"""The three export formats, over one row shape.

Moved here with `core/exports/tabular.py` itself when `examinations` became its
second caller. The cases are unchanged: they were always about the renderer, not
about attendance — which is precisely the property that let the module move.

`SpreadsheetInjectionTests` is the regression suite for a real finding. A remark
a teacher typed must not execute when someone opens the export, and openpyxl
writes a leading `=` as a genuine formula, so the guard is not only defence
against the reader's spreadsheet — it stops us authoring one ourselves.
"""

from __future__ import annotations

import datetime
import io

from django.test import SimpleTestCase

from core.exports import tabular

ROWS = [
    {"admission_number": "S-1", "first_name": "Ayesha", "attendance_rate": "92.5"},
    {"admission_number": "S-2", "first_name": "Bilal & Co", "attendance_rate": "40.0"},
]


class ExportFormatTests(SimpleTestCase):
    """One row shape, three renderings — the point of a formatter-only split."""

    def test_csv_carries_the_headers_and_every_row(self) -> None:
        data, mime_type, extension = tabular.render(ROWS, fmt="csv", title="Summary")

        text = data.decode()
        self.assertEqual((mime_type, extension), ("text/csv", "csv"))
        self.assertIn("admission_number,first_name,attendance_rate", text)
        self.assertIn("Bilal & Co", text)

    def test_csv_says_so_when_there_is_nothing_to_say(self) -> None:
        """An empty file is indistinguishable from a failed export when someone
        opens it."""
        data, _, _ = tabular.render([], fmt="csv", title="Summary")

        self.assertIn("no rows matched", data.decode())

    def test_xlsx_is_a_real_workbook_with_a_frozen_header(self) -> None:
        import openpyxl

        data, mime_type, extension = tabular.render(ROWS, fmt="xlsx", title="Summary")

        self.assertEqual(extension, "xlsx")
        self.assertIn("spreadsheetml", mime_type)
        workbook = openpyxl.load_workbook(io.BytesIO(data))
        sheet = workbook.active
        self.assertEqual([cell.value for cell in sheet[1]], list(ROWS[0]))
        self.assertEqual(sheet["B3"].value, "Bilal & Co")
        self.assertEqual(sheet.freeze_panes, "A2")

    def test_a_sheet_name_excel_would_reject_is_sanitised(self) -> None:
        """Excel refuses a name over 31 characters or containing []:*?/\\ and
        fails the whole save rather than truncating."""
        import openpyxl

        data, _, _ = tabular.render(
            ROWS, fmt="xlsx", title="Daily register: 6-A / 2026 [draft] " * 3
        )

        sheet = openpyxl.load_workbook(io.BytesIO(data)).active
        self.assertLessEqual(len(sheet.title), 31)
        self.assertNotIn(":", sheet.title)

    def test_pdf_renders_and_escapes_its_values(self) -> None:
        data, mime_type, extension = tabular.render(ROWS, fmt="pdf", title="Summary")

        self.assertEqual((mime_type, extension), ("application/pdf", "pdf"))
        self.assertTrue(data.startswith(b"%PDF"))

    def test_pdf_refuses_a_row_count_nobody_would_read(self) -> None:
        """A 40,000-row register is hundreds of pages and enough memory to
        matter. The caller is told, rather than handed a truncated document that
        looks complete."""
        too_many = [dict(ROWS[0]) for _ in range(tabular.PDF_ROW_LIMIT + 1)]

        with self.assertRaises(tabular.ReportTooLargeForFormat):
            tabular.render(too_many, fmt="pdf", title="Summary")

    def test_an_unknown_format_is_refused(self) -> None:
        with self.assertRaises(ValueError):
            tabular.render(ROWS, fmt="docx", title="Summary")

    def test_a_date_is_rendered_iso_not_locale_formatted(self) -> None:
        """A locale-formatted date is the classic way a column of dates becomes
        a column of text in a spreadsheet."""
        rows = [{"attendance_date": datetime.date(2026, 4, 6)}]

        data, _, _ = tabular.render(rows, fmt="csv", title="Register")

        self.assertIn("2026-04-06", data.decode())


class SpreadsheetInjectionTests(SimpleTestCase):
    """A remark a teacher typed must not execute when the export is opened.

    `remarks` is free text on the register, so this vector starts inside our own
    data and needs no other flaw to reach a reader's Excel.
    """

    DANGEROUS = [
        {"remarks": '=HYPERLINK("http://evil","Click")'},
        {"remarks": "+1+1"},
        {"remarks": "-1+1"},
        {"remarks": "@SUM(A1:A9)"},
    ]

    def test_csv_neutralises_every_formula_trigger(self) -> None:
        data, _, _ = tabular.render(self.DANGEROUS, fmt="csv", title="Register")

        text = data.decode()
        for line in text.splitlines()[1:]:
            self.assertTrue(
                line.lstrip('"').startswith("'"),
                f"a formula trigger reached the cell unescaped: {line}",
            )

    def test_xlsx_writes_text_not_a_formula(self) -> None:
        """openpyxl writes a string beginning `=` as a *formula*, so this is not
        only defence against the reader's spreadsheet — it stops us authoring
        one ourselves."""
        import openpyxl

        data, _, _ = tabular.render(self.DANGEROUS, fmt="xlsx", title="Register")

        sheet = openpyxl.load_workbook(io.BytesIO(data)).active
        values = [row[0].value for row in sheet.iter_rows(min_row=2)]
        self.assertTrue(all(value.startswith("'") for value in values), values)

    def test_an_ordinary_remark_is_left_alone(self) -> None:
        """The control: the guard must not be prefixing everything."""
        data, _, _ = tabular.render(
            [{"remarks": "Left early for a dental appointment"}], fmt="csv", title="R"
        )

        self.assertNotIn("'Left early", data.decode())

    def test_the_pdf_is_not_prefixed(self) -> None:
        """There is no formula engine in a PDF, and a stray apostrophe in a
        printed register would be a defect with nothing to justify it."""
        data, _, _ = tabular.render(self.DANGEROUS, fmt="pdf", title="Register")

        self.assertTrue(data.startswith(b"%PDF"))
